# -*- coding: utf-8 -*-
"""
Created on Sun Apr 11 14:19:00 2021

@author: ntruo
"""

import cv2
import requests

# initialize the Keras REST API endpoint URL along with the input
# image path
KERAS_REST_API_URL = "http://127.0.0.1:5000/predict"

import io
import numpy as np
import base64
from PIL import Image
import cv2
import pandas
import os
from openpyxl import load_workbook
import xlsxwriter
#%%
# def stringToImage(base64_string):
#     imgdata = base64.b64decode(base64_string)
#     return Image.open(io.BytesIO(imgdata))

# # convert PIL Image to an RGB image( technically a numpy array ) that's compatible with opencv
# def toRGB(image):
#     return cv2.cvtColor(np.array(image), cv2.COLOR_BGR2RGB)

# def unpack_img(img):
#     img = "000b'/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCADnAL8DASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwDa9s0vTvXG/bbnp58n/fRpPtlx3nk/76NeY80X8p1rAeZ2YfsadvGK4r7VcZz50n/fRoa6nPWWQfRjS/tRfyj+o+Z2e7tSFgDkGuN+0zf89ZP++jTfOl6ea/8A30aP7U/uh9Q8ztdw9eaUSY/irifOk6+Y/wCLGl86TH+sb86TzT+6NYFLqdsZVI+9TCy45Iri97n+M/nSb2P8R/Ol/aj/AJQeBT6nab0H8a/nQZUyPnX864os3difxoDY7mj+1Zfyh9Qj3O081O7r+dIZoR/y0TP1rjNwpNwpf2pP+UPqEe52n2iLP+tT86UXcCjmZP8AvquKzRupPNJ/ylLAxXU7U3luf+W0f/fQphu7bJzNH/31XG7qTdj0pf2pPsN4KD6nYm9tQP8AXx/99UDULVes6fnXH7ufajNJ5pU7AsDDude2o2h5+0R/nTf7Rs+9wh/GuSzSZo/tSr2QfUKZ139pWY63CfnU0N/bTv5ccqs3XAriya0dFONRBzj5DV0swqTmotEzwkIRbM3PbFL6UmOlHNeWdyF5zS8+tNpaQBRjmik70DFIo6UlNdwq5JoWomxxIFNMqisDVPEUdoxjiG9/0Fc3Prl5cEnzSo9F4ruo4GpNXeiOapiYx0R373UMfWRR9TikW7jYkB1J9Aa80a4kc5d2b6mhJ5Iz8jFc+ldP9mK3xGP1x9j07zl9RTgynoa8yN7cf893/M1atdbvLZwwlL+zc1EstlbRlLFq+qPRRyOBS9a53T/E0M4VJyUf9K3IrmOYAq2QfeuGpQnTdpI6oVYy2JufSjB9KTIpc1jY0DBpccUhIozSGGKKKKADFaWiLnUB3+U1m1paIcakp/2WrbD/AMWJlW+BmXjnI6U7NN6GlBrMtC0c0ZxQDg9aQC4PoRTT19qUk561XuZ1gjZ2PA9aqMXJ2QpSsrjLu8itIjJI4UCuU1TxFJP8luSqnqe9UdUv5r+4YFv3YPAqtHa56mvaw+DjBKU9zzateUnZbFZt7sWY5J705Y8960EsRj19qmSwYED5vwFd9+xz2MoRZ7n8qcIcHnNbg0o8EK5/Cp10Z8jKYB707NhdHP8A2Z8DIOKYbc46Gu6j0SNo15A454qnd6Fj/Vn5e/FVySJ5onH+WR0NXLK/uLOVWWRio6rmprixaIkHpVR4tvHSspxUlaSLi2tUd3p+oxXsQZW5xyKvjmvPbG8ls5g4c7c8iu10+/jvIAyt+FeJisK6bvHY9GjW5tGXsUtJ+NFcJ1IXvSmm5o60rDF71o6MM6iv+6aze9aOi4GpJ/umtsP/ABYmVb4GZtAFFGPesyhaB78UlLk96AEYhQcmuU8QajuJt43x/e4ro7yURW7tkDAzzXBSsZZnlfksc16OApJvmfQ48TO3uojRABktyasRRM7DB/So0+ZsAZY9K6bStMKAPIuXPPsK9qMHJnC5JFax0yV3HBI689q6CKwVY/ljGfUir1vb7FHQ1a8on0rqjTSOeU2zLWHYM4qWOJ3PKhV9+9aC24784qURgDpWmhDZS8jAqncxuo3AZGelbRXjoKgmi3rjGKZNzk9QtS6blHJ9qwJYTtOFP4iu6mgAUqV6frXN6jbBMuowp71jVpdUdFOfQ5uWI9VBFW9IvZLW7VQ3yMeRmqsxwxHp7VVJZHV1HQ5rjnBSi4s3jKzuenxNuQH2qTFZGiXpu7JWJG4da1RzXzVWDhJxZ69OXNFMdij8ab3pf5VmWLj3rR0Yf8TFTkfdas3FaOiDOpoP9lq2w/8AFiZ1vgZnd6BTeaUcVkUO70GkpKLAzG166ENow/vcVxxbLcHit3xTKcogHvXPQAvIBivfwdPlpJ9zy68rzOi0PTxcSCRhwOldtb2yxqABWToFt5dupIxx0rokXAr1aasjhqO7EVBUwWkAqRVqyAC9qXbUgFLxSuOxAVpjLkVO+1epwKjxkcU0xNFOWLeKyr2yWVSjDrW6y+1QSxBwciruLVHmmq2ZtWJHIBwax34JUn5T0rvdd0tXiZ1XPHI9a4GaPynKdhXLVjZnTCV0dH4UlXfJHu+gNdeD7159oEpi1KPnAPWu/Q/KD7V89mELVL9z1MLK8LD6OM0DrSivPOsQGtHRf+QkgHUq1Z9aGj8agv8AumtsP/FiZ1vgZmmjNJQKzGKDSMcA0ue1MlcKhJppaik9Di/E0pN2iegqnpVsZ7lQOuaNbnW41NyvQcZro/DGnhYfOZeT0r6bDQtCKPIqy1bOmsIRFCi+grTXAFY11qEVioGGZyOFXrVeLUb2b5ktmK+hOK7W7HMot6nRjrUo6VgpeXqH54cDvVu31FiQJExnpjpSvcOWxrgZ6U7bVeO5Vqn80UCEZAeozTGAApZJsDgVn3F+V4VCTTAtMaiJBrOe5vJCQi8ewqlO2pIpIQ/pT5h8ho3cYkQrivOdesjbXrMFAVuhzXYWmrvLOLe4iZHPQmszxTaloN4GSpqaivEcLxdjkrJ/LuUbHQ+tei2riSBGHII9a80UNv6c13uhbjp6Z/CvEzGHuqR6WElq0ao+lLnHakApQOK8Y9AMnNaOjEf2iuf7prOArQ0dc6gvGeDWuH/ixM6vwMziaAaQ5pRnHNZlC1l6rLKICkYJZuK0yeDWLql21rcwOQShbDYHSurBxUqqTMMQ7QZyE0YW9aPH3SAfrXoukQhbGMKMcVyM1hi7eVuQ8m7P413mnqotoxjtX00I2Z5E3oQrZoJmkcZY9z2q2rRR8cD2ouQ235OtYJ066lui8944jJ/1a8D862sZbnQiaFuNwo2RkcYxXByaPq66kVi3bC3EivwBmuut4pbbYrztJwASR3qU7lONjQXAqYE4qAqVYc9asoflp2JIZSTxnFQFVzk/rU0py2B1qlcL5rNCJGTjlgKdhbkrXUEIwzAUwXkEwwHB9K4ddG1JdR8tgzLu/wBYzZBHrW7c6WolV7SR0dcfdPBojqU4ruak1pFM4bHI5BFZ2vRFtPl+UsQuRWraRyCMCU5OKg1WNXtJAfSnJaBF6nA6PYQ3sk8UyHftyuDjFdXocflWSoOQpIGaxbVY9N+0XpDHCED3PYVuaIrf2dEX+8eTmvFzNWpJHo4T42aXWlAPbFGeMUGvBPTFw3pV/Rh/xMY+ezfyqhV/R/8AkIp67TWuH/ixM6vwMzCaAfam5pfpWZYuRisbXYlksyccqcitis7Vl3WUgyehrfDytUTMqyvBkbur2VvI6DBA/lXRWjZRcelc9ZJ9t0WLaBlV28+ora01/wBymeuK+rieHLY1du4DgUeQpHKj8afHjFSgVRKRXW3RTkIM1IY8DNTAU1xxQMrEHPPalB4zTXOKVTxVkEZzvqQLn0qInDVOoJTihghv2dW5IB/Cj7Oo6AflVlFwoBpSO1SmUU2j2jisrVGxbuCccVtyAYNc7rKSzBYofvM4H60witSr9hjuNNijmAILZIx171dgiWCIIvQVJJH5RSIDG1efem4yK+YzGq5VnG+iPawkLQuOpQec800UorzjrFGO+a0NHwNRT6HH5Vnd6v6Of+Jin0P8q2ofxImdX4GZntSUUdBUFC1BdRebbso6kelTZ70GnFuLuhNXVjH8OyG3ebT5iAwbdHk9a3LI+W7Rkg7WrI1HTFuwGUlXHRlODSaDDc20kkc7MwHILHJr6LC4yNRJPc8mth3Bt9DsI2yKsqaowvwKsK/vXoHIWN1RzOACc0m/3qOXLKRQJsq+YXb0FToPlqt+8HAUcVKk5VSDWhNiObcrZFWbZ9yZNVXZ2OQvHvVi3GxcHjND2BF0HikJpobims4qCiOVsA1lRyRteszMBsHQmr078GuUk01pNZa8dsqOAM1z4nERowuzehSdSVjaml8yZn5x0H0plIOgpa+TqTc5OT6nuQiopJAKWkpagsM1oaOCdRTHBwf5Vn1oaOcalH9D/KtaH8SJnV+BmVS0hOTRUlBRRmkoAMA0IQjg/nRUU7FYyR1rWjJxmmjOpFSi0bMTZAxVhTxXOaLq63W+FziSPjHqK31fivq6cuaNzwZqzsTbsUwuKaTVO6eUD93gfWtEQW2YD0qMyDNZrJcsdwk3e3SgRXLfwt+dWolcpoiUVIkoNY5hucEfc9yc1Lai4B+aTcvanYTibIkoLcZquh45pWcAGpZJFcSbVJNUAB17mnTziWXapyq9T6mmg181mVf2lTlWyPawdLlhfuOopAaWvMO0KWkopDHdutX9I/5CUYHof5Vn1e0c/wDEyj9MH+Va0P4kTOr8DMyikzS9qQwoopKAAcGorgZiI9qlBqKY4RvpVR3QpbHGJLPZ6yhgLbi/QdxmvRLa48yMZ4bHIrj9PULqfmuozMzBOOgHeupjj/drg4Yd6+ow7vFHiVbXNENmkIB4Iqos5QgScVaRwwFdBgMMTA5TFH+kEcgcVYBFLuFUmFyo0Ukh+cgD2p6oEGBUxNRuwVTk07ivcbuxWF4j1Se0smFsjM/8TD+AetaMkxkbbGDju1RR28cjvE670kUqwPeonsVBamLoFz5tqVZizZzknmtsc1y2nI2m69c2LZ2ox2+69R+mK6cHivlsZDlqvzPcw8rwQ8UZpKM1yHQOzSimUuaLAOq9pBxqKfQ/yqhV/R8nUo8DPDcfhWlD+JEzq/AzMopuaNxpFDjSUUnIpALSMhfgDk0oyxwAST2FaFlp0ryq0qYX07mtaVKU5KyInNRWpmazYLCluUbDQLn6+tXLVt8KN6gGoPGCyReQ69GJUinaex+zoD1Ar6WhGyPFrO5beIOORUBE0XKHI9DVwdKCoNdJgVlu5B9+Ij6c0G9/6ZufwqcxnPFKIz6D86AK5u3b7sT/AI8VGElnP7w4X+6KuCPmn7QKLgVxEFGAMVQe6FtqkQYZD8DHrWpIcA1lR2Y1DWYgTxH85A/SlPWJVPcz/FOnywa1a38ETMsg2yFRnB7ZrSjOYwa6mWBXUAryBVVtOgm4Zdreq8V4uMw8qtnE9LD1FBWZh9qTNX7nSpoiTGC6/rVFlKk7gQfcYryZ05QdpI74zUloFFJ+FFZljq0NIIGoxn2P8qzhV/ScHUYx04P8q1o/xEZ1fgZl+9A96QDpVq2sZ7ggquE/vEURhKTskEpKOrK/XFXINMuLkBsBFPdq1bTTorf5gpZ/7xrRRcV6FHA9ZnNPEdIlK102O3j+UfN6nrVuOPDZNTYHpRXoRgoqyOVtvVmZrNit5aFCAe49qwrZDGAp6jiutkXcpFYd3bhJd6jGetdNKXQ56i6jFNP6ioxTxW5gOpaSloAKTOKU000ARy8irmj2iqXmIwScZ9qqhNzAetblvGIogo9KyqSsrGtKN2D8MaZjnNK/+sNAPNcx0EqAMOetI8EbjlAfqKReKkBqXFPcpNozrjSYJBlF2N7VlT6Xcw8hd6+q10wxTsDFctTB057aG0K8onGFGQgMrAn1FXdKP/ExjP1/ka37i2jnQq6g5GKzbbTZLa+R1IZBnn04rkeDnComtUbOupQdx9vpkEBDFdzdctV9EA6dKUjmmZKmvVhTjBWijjcnLckI4opA2ec0vU1Yhc0tJSE0gHHFUrqPepqw74HSmkF0+br7VSdmKSujFYYOKUVNcQsrlgCRUArqi7o42rMeOlKOab1GD0py4C4HaqEONNIpC1CK0jgDqaTGixZxb5cnoK1hwcVXt4xEmBU3UH3rlnK7OqEbIif75oWmhSvBzj1NOBrMsevSn0xafTAUVIDxUeaN2elIBxPpSqPmFNAxyacOtNA9iEHijg0UVQhgHPHSpAaKKQBmmk4HNFFAxmM8n8KQSqSV5yKKKEDGlAWwapXNt5bbl6GiitIOzMZpWK3Q0vJooroMBQM1oWkIT5iOaKKyqPQ0prUtgc0jy7SEUZP8qKK5zpQpG4YNREYoopASKe1OoooAKVRzRRQA7PalBwRmiimDP//Z'"
#     imgdata = base64.b64decode(img)
#     filename = 'some_image.jpg'  # I assume you have a way of picking unique filenames
#     with open(filename, 'wb') as f:
#             f.write(imgdata)

# def unpack1(img):
#     # imgdata = base64.b64decode(img)
#     input_ = np.fromstring(img, np.uint8)
#     input_ = cv2.imdecode(input_,cv2.IMREAD_COLOR)
#     cv2.imshow("x", input_)

#%%

def to_excel(sample_dict):
    headers = ["Name", "DOB", "Faculty","Admission Year", "Bank ID","Student ID"]
    book = load_workbook('info_extracted.xlsx')
    sheet = book['Sheet1']
    values = [sample_dict[key] for key in headers]
    sheet.append(values)
    book.save(filename='info_extracted.xlsx')
    book.close()


#%%
files = {
    'input_': open('2.jpg', 'rb')
}

response = requests.post(KERAS_REST_API_URL,  files=files).json()

if response["success"]:
    print(response['predictions'])
    # to_excel(response['predictions'])
else:
    print(response['predictions'])
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
#%%

import xlsxwriter


 
item2 = {'Name': 'Nguyen Truong An',
 'DOB': '11/05/2000',
 'Faculty': 'Công nghệ thông tin',
 'Admission Year': '2018',
 'Bank ID': '9704 1800 9364 1124',
 'Student ID': '18110246'}

collumn_name = ["Name","DOB","Faculty","Admission Year","Bank ID","Student ID"]

header = {'Name': 'Name',
 'DOB': 'DOB',
 'Faculty': 'Faculty',
 'Admission Year': 'Admission Year',
 'Bank ID': 'Bank ID',
 'Student ID': 'Student ID'}

# rows = [header,item1,item2]
#%%

#%%

#%%
item1 = {'Name': 'Trường ',
 'DOB': '29/05/2000',
 'Faculty': 'Công nghệ thông tin',
 'Admission Year': '2018',
 'Bank ID': '9704 1800 9364 1124',
 'Student ID': '18110255'}
sample_dict = item1

if __name__ == '__main__':
    headers = sample_dict.keys()
    book = load_workbook('info_extracted.xlsx')
    sheet = book['Sheet1']
    values = [sample_dict[key] for key in headers]
    sheet.append(values)
    book.save(filename='info_extracted.xlsx')
    book.close()



